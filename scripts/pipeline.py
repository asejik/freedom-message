"""
pipeline.py — CLC Sermon Platform: Data Consolidation Pipeline
==============================================================
Epic 3, Step 1: Extract → Match → Normalize → Serialize

What this script does (NO network calls, NO API calls):
  1. Reads all 12 sheets of output/CLC_Sermons.xlsx into a flat record list.
  2. For each row, locates the matching transcript .txt in transcripts/{YEAR} transcripts/
     using the S/N prefix as the primary key (e.g., "7_Title.txt" matches S/N=7).
  3. Normalizes all fields to match the Supabase schema defined in web/supabase/schema.sql.
  4. Serializes the unified dataset to output/clc_consolidated_staging.json.

Output shape (per record):
  {
    "sn":               int,        # original row number within the year sheet
    "year":             int,
    "title":            str,
    "series":           str | null, # maps to series.name
    "preacher":         str | null, # maps to preachers.name
    "date_preached":    str,        # ISO 8601 "YYYY-MM-DD"
    "audio_url":        str,
    "transcript_text":  str | null, # full .txt content, null if file not found
    "transcript_file":  str | null, # relative path of matched file
    "transcript_matched": bool,
    "ai_summary":       null,       # placeholder — populated in Groq enrichment step
    "ai_tags":          []          # placeholder — populated in Groq enrichment step
  }
"""

import os
import re
import json
import glob
import openpyxl
from datetime import datetime

# ── Paths ─────────────────────────────────────────────────────────────────────
SCRIPT_DIR      = os.path.dirname(os.path.abspath(__file__))
ROOT            = os.path.join(SCRIPT_DIR, '..')
XLSX_PATH       = os.path.join(ROOT, 'output', 'CLC_Sermons.xlsx')
TRANSCRIPTS_DIR = os.path.join(ROOT, 'transcripts')
OUTPUT_PATH     = os.path.join(ROOT, 'output', 'clc_consolidated_staging.json')

YEARS = list(range(2015, 2027))

# ── Date Parsing ──────────────────────────────────────────────────────────────
MONTH_MAP = {
    'january': 1, 'february': 2, 'march': 3, 'april': 4,
    'may': 5, 'june': 6, 'july': 7, 'august': 8,
    'september': 9, 'october': 10, 'november': 11, 'december': 12
}

def parse_date(raw: str | None) -> str | None:
    """Convert 'DD Month YYYY' → 'YYYY-MM-DD'. Returns None on failure."""
    if not raw:
        return None
    try:
        parts = str(raw).strip().split()
        if len(parts) == 3:
            day   = int(parts[0])
            month = MONTH_MAP.get(parts[1].lower())
            year  = int(parts[2])
            if month:
                return datetime(year, month, day).strftime('%Y-%m-%d')
    except (ValueError, IndexError):
        pass
    # Fallback: try Python's parser
    try:
        return datetime.strptime(str(raw).strip(), '%d %B %Y').strftime('%Y-%m-%d')
    except ValueError:
        pass
    return None


# ── Transcript Matching ───────────────────────────────────────────────────────
def build_transcript_index(year: int) -> dict[int, tuple[str, str]]:
    """
    Build a mapping of {sn: (abs_path, rel_path)} for all .txt files
    in 'transcripts/{YEAR} transcripts/'.

    Filename convention: '{S/N}_{Title}.txt'
    The S/N prefix is the integer before the first underscore.
    """
    folder = os.path.join(TRANSCRIPTS_DIR, f'{year} transcripts')
    index: dict[int, tuple[str, str]] = {}

    if not os.path.isdir(folder):
        return index

    for filepath in glob.glob(os.path.join(folder, '*.txt')):
        filename = os.path.basename(filepath)
        # Extract the leading integer (S/N)
        m = re.match(r'^(\d+)_', filename)
        if m:
            sn = int(m.group(1))
            rel = os.path.relpath(filepath, ROOT)
            index[sn] = (filepath, rel)

    return index


def read_transcript(abs_path: str) -> str:
    """Read and return the full text content of a transcript file."""
    try:
        with open(abs_path, 'r', encoding='utf-8', errors='replace') as f:
            return f.read().strip()
    except OSError:
        return ''


# ── Main Pipeline ─────────────────────────────────────────────────────────────
def run():
    print('=' * 60)
    print('  CLC Sermon Platform — Data Consolidation Pipeline')
    print('=' * 60)

    print(f'\nLoading workbook: {XLSX_PATH}')
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True)

    records         = []
    year_stats      = {}          # {year: {'total': n, 'matched': n, 'missing': [titles]}}
    all_preachers   = set()
    all_series      = set()
    total_matched   = 0
    total_missing   = 0

    for year in YEARS:
        sheet_name = str(year)
        if sheet_name not in wb.sheetnames:
            print(f'  [SKIP] Sheet "{sheet_name}" not found in workbook.')
            continue

        ws    = wb[sheet_name]
        index = build_transcript_index(year)

        year_total   = 0
        year_matched = 0
        year_missing = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            # Columns: S/N, Title, Series, Preacher, Date, URL
            sn, title, series, preacher, date_raw, audio_url = (
                row[0], row[1], row[2], row[3], row[4], row[5]
            )

            # Skip completely empty rows
            if not title and not audio_url:
                continue

            year_total += 1
            sn_int = int(sn) if sn is not None else year_total

            # ── Transcript matching ──────────────────────────────────────
            transcript_text    = None
            transcript_file    = None
            transcript_matched = False

            if sn_int in index:
                abs_path, rel_path = index[sn_int]
                content = read_transcript(abs_path)
                if content:
                    transcript_text    = content
                    transcript_file    = rel_path
                    transcript_matched = True
                    year_matched      += 1
                    total_matched     += 1
                else:
                    year_missing.append(f'S/N {sn_int}: {title} (empty file)')
                    total_missing += 1
            else:
                year_missing.append(f'S/N {sn_int}: {title}')
                total_missing += 1

            # ── Field normalization ──────────────────────────────────────
            clean_preacher = str(preacher).strip() if preacher else None
            clean_series   = str(series).strip()   if series   else None
            clean_title    = str(title).strip()     if title    else ''
            clean_url      = str(audio_url).strip() if audio_url else ''
            date_iso       = parse_date(date_raw)

            if clean_preacher:
                all_preachers.add(clean_preacher)
            if clean_series:
                all_series.add(clean_series)

            records.append({
                'sn':                sn_int,
                'year':              year,
                'title':             clean_title,
                'series':            clean_series,
                'preacher':          clean_preacher,
                'date_preached':     date_iso,
                'audio_url':         clean_url,
                'transcript_text':   transcript_text,
                'transcript_file':   transcript_file,
                'transcript_matched': transcript_matched,
                'ai_summary':        None,   # Groq enrichment step
                'ai_tags':           [],     # Groq enrichment step
            })

        year_stats[year] = {
            'total':   year_total,
            'matched': year_matched,
            'missing': year_missing,
        }

        match_rate = (year_matched / year_total * 100) if year_total else 0
        print(f'  [{year}]  {year_total:3d} rows  |  '
              f'{year_matched:3d} transcripts matched  |  '
              f'{year_total - year_matched:2d} missing  |  '
              f'{match_rate:.0f}% match rate')

    wb.close()

    # ── Unique reference sets ─────────────────────────────────────────────────
    preachers_list = sorted(all_preachers)
    series_list    = sorted(all_series)

    # ── Serialize to JSON ─────────────────────────────────────────────────────
    payload = {
        'metadata': {
            'generated_at':    datetime.utcnow().isoformat() + 'Z',
            'total_records':   len(records),
            'total_matched':   total_matched,
            'total_missing':   total_missing,
            'years_processed': YEARS,
            'unique_preachers_count': len(preachers_list),
            'unique_series_count':    len(series_list),
        },
        'unique_preachers': preachers_list,
        'unique_series':    series_list,
        'records':          records,
    }

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    # ── Final summary ─────────────────────────────────────────────────────────
    print()
    print('=' * 60)
    print(f'  Total records:          {len(records)}')
    print(f'  Transcripts matched:    {total_matched}  ({total_matched/len(records)*100:.1f}%)')
    print(f'  Missing transcripts:    {total_missing}')
    print(f'  Unique preachers:       {len(preachers_list)}')
    print(f'  Unique series:          {len(series_list)}')
    print(f'  Output saved to:        {OUTPUT_PATH}')
    print('=' * 60)

    if total_missing > 0:
        print('\n  Missing transcript details:')
        for year, stats in year_stats.items():
            if stats['missing']:
                print(f'\n  [{year}]')
                for m in stats['missing']:
                    print(f'    • {m}')

    print('\n✅ Pipeline complete. Ready for Groq enrichment step.\n')


if __name__ == '__main__':
    run()
