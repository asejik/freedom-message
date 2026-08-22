"""
normalize_preachers.py
Cleans up typos and inconsistent preacher names in CLC_Sermons.xlsx,
then lists all unique preachers after the fix.
"""

import os
import openpyxl

SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(SCRIPT_DIR, '..', 'output', 'CLC_Sermons.xlsx')

# ── Canonical name mapping ────────────────────────────────────────────────────
# key = exactly what's in the file  →  value = corrected form
NORMALIZATIONS = {

    # ── Apostle Muyiwa Areo (typos + Pastor → Apostle upgrade) ──
    'Apostle  Muyiwa Areo':     'Apostle Muyiwa Areo',   # double space
    'Apostle Muiywa Areo':      'Apostle Muyiwa Areo',   # transposed letters
    'Apostle Muyiwa Aeo':       'Apostle Muyiwa Areo',   # missing r
    'Apostle Muyiwa Areo;':     'Apostle Muyiwa Areo',   # trailing semicolon
    'Pastor Muyiwa Areo':       'Apostle Muyiwa Areo',   # wrong title
    'Pastor Muyiwa Are':        'Apostle Muyiwa Areo',   # wrong title + typo
    'Pastor muyiwa Are':        'Apostle Muyiwa Areo',   # lowercase + typo
    'PastorMuyiwa Areo':        'Apostle Muyiwa Areo',   # missing space

    # ── Apostle Segun Obadje (typo + Pastor → Apostle upgrade) ──
    'Apoastle Segun Obadje':    'Apostle Segun Obadje',  # typo
    'Pastor Segun Obadje':      'Apostle Segun Obadje',  # wrong title

    # ── Pastor Temitope Areo (typos) ──
    'Pastor Temitope Areo;':    'Pastor Temitope Areo',  # trailing semicolon
    'Pastor Temitope  Areo':    'Pastor Temitope Areo',  # double space
    'Pastor Temtope Areo':      'Pastor Temitope Areo',  # missing i
    'Pastor Temitpe Areo':      'Pastor Temitope Areo',  # missing o
    'Pastor temitotpe Areo':    'Pastor Temitope Areo',  # lowercase + typo
    'Pastor Temi Areo':         'Pastor Temitope Areo',  # nickname → full name

    # ── Evangelist Chukwuka Okoye (typos) ──
    'Evan. Chuks Okoye':        'Evangelist Chukwuka Okoye',
    'Evangeslist Chukwuka Okoye': 'Evangelist Chukwuka Okoye',

    # ── Pastor Israel Omotade ──
    'Pastor Isreal Omotade':    'Pastor Israel Omotade',

    # ── Pastor Damilola Faleye ──
    'Pastor Dami Faleye':       'Pastor Damilola Faleye',

    # ── Rev. Kayode Ijisesan ──
    'Pastor Kayode Ijisesan':   'Rev. Kayode Ijisesan',

    # ── Pastor Funke Obadje ──
    'Pastor Mrs Funke Obadje':  'Pastor Funke Obadje',
}

# ── Load workbook ─────────────────────────────────────────────────────────────
print(f"Loading {OUTPUT_FILE} ...")
wb = openpyxl.load_workbook(OUTPUT_FILE)

fixed_count  = 0
sheet_counts = {}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    sheet_fixed = 0
    for row in ws.iter_rows(min_row=2):
        cell = row[3]   # column D = Preacher
        if cell.value and str(cell.value).strip() in NORMALIZATIONS:
            old = cell.value
            cell.value = NORMALIZATIONS[str(cell.value).strip()]
            sheet_fixed += 1
            fixed_count  += 1
    if sheet_fixed:
        sheet_counts[sheet_name] = sheet_fixed

print(f"\nFixed {fixed_count} cells across {len(sheet_counts)} sheets:")
for sheet, n in sheet_counts.items():
    print(f"  {sheet}: {n} fix(es)")

# ── Save ─────────────────────────────────────────────────────────────────────
wb.save(OUTPUT_FILE)
print(f"\n✅ Saved: {OUTPUT_FILE}")

# ── List all unique preachers after cleanup ───────────────────────────────────
print("\n─── Unique Preachers (post-cleanup) ───────────────────────────────")
wb2 = openpyxl.load_workbook(OUTPUT_FILE, read_only=True)
preachers = set()
for sheet_name in wb2.sheetnames:
    ws = wb2[sheet_name]
    for row in ws.iter_rows(min_row=2, values_only=True):
        p = row[3]
        if p and str(p).strip():
            preachers.add(str(p).strip())
wb2.close()

print(f"Total unique: {len(preachers)}\n")
for p in sorted(preachers):
    print(f"  {p}")
