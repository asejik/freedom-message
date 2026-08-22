import csv
import re
import io
import os
import requests
from mutagen.id3 import ID3, ID3NoHeaderError
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

YEARS = [2015, 2016, 2017, 2018, 2019, 2020, 2021, 2022, 2023, 2024, 2025, 2026]
HEADERS = ['S/N', 'Title', 'Album', 'Preacher', 'Date', 'URL']

# --- Styling ---
HEADER_FILL   = PatternFill("solid", fgColor="1F3864")   # dark navy
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
CELL_FONT     = Font(name="Calibri", size=10)
ALT_FILL      = PatternFill("solid", fgColor="EEF2F7")   # light grey-blue
THIN_BORDER   = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

COL_WIDTHS = {
    'S/N':      6,
    'Title':    60,
    'Album':    40,
    'Preacher': 28,
    'Date':     18,
    'URL':      80,
}


def extract_date(title):
    """Parse (DD-MM-YY) or (DD-MM-YYYY) with optional a/b suffix from title."""
    match = re.search(r'\((\d{1,2})-(\d{1,2})-(\d{2,4})[ab]?\)', title)
    if match:
        day, month, year = match.groups()
        if len(year) == 2:
            year = '20' + year
        try:
            return datetime(int(year), int(month), int(day)).strftime('%d %B %Y')
        except ValueError:
            pass
    return ''


def get_metadata(url):
    """Fetch the first 128 KB of the MP3 and read ID3 tags."""
    try:
        resp = requests.get(url, headers={'Range': 'bytes=0-131071'}, timeout=20)
        tags = ID3(io.BytesIO(resp.content))
        preacher = str(tags.get('TPE1', '')).strip()
        album    = str(tags.get('TALB', '')).strip()
        return preacher, album
    except Exception:
        return '', ''


def style_sheet(ws, num_rows):
    """Apply formatting to a worksheet."""
    # Header row
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[header]

    ws.row_dimensions[1].height = 22

    # Data rows
    for row_idx in range(2, num_rows + 2):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font   = CELL_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                vertical='center',
                wrap_text=(col_idx == 2),          # wrap Title column only
                horizontal='left' if col_idx != 1 else 'center'
            )
            if fill:
                cell.fill = fill

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"


# ── Main ──────────────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)  # remove default blank sheet

for year in YEARS:
    csv_file = f'sermon_urls_{year}.csv'
    if not os.path.exists(csv_file):
        print(f"[SKIP] {csv_file} not found.")
        continue

    # Read CSV
    rows = []
    with open(csv_file, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        next(reader)   # skip header
        for row in reader:
            if len(row) >= 2 and row[1].strip():
                rows.append((row[0].strip(), row[1].strip()))

    total = len(rows)
    print(f"\n[{year}] Fetching metadata for {total} sermons...")

    # Fetch metadata concurrently (10 workers)
    metadata = {}
    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = {executor.submit(get_metadata, url): idx for idx, (_, url) in enumerate(rows)}
        done = 0
        for future in as_completed(futures):
            idx = futures[future]
            metadata[idx] = future.result()
            done += 1
            if done % 20 == 0 or done == total:
                print(f"  {done}/{total} done", flush=True)

    # Create sheet
    ws = wb.create_sheet(title=str(year))
    ws.append(HEADERS)

    for sn, (title, url) in enumerate(rows, start=1):
        preacher, album = metadata.get(sn - 1, ('', ''))
        date = extract_date(title)
        ws.append([sn, title, album, preacher, date, url])

    style_sheet(ws, total)
    print(f"  Sheet '{year}' created ({total} rows).")

output_file = 'CLC_Sermons.xlsx'
wb.save(output_file)
print(f"\n✅ Saved: {output_file}")
